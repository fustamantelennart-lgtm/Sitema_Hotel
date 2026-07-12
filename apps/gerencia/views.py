from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from apps.usuarios.decorators import rol_requerido
from .services import GerenciaService
import json


@login_required
@rol_requerido('admin')
def dashboard(request):
    from apps.recepcion.models import Hotel, Habitacion

    hotel  = Hotel.objects.first()
    kpis   = GerenciaService.get_kpis(hotel)
    meses_labels, meses_ocupacion = GerenciaService.get_grafico_reservas()

    from apps.recepcion.models import Habitacion
    habitaciones  = Habitacion.objects.all()
    disponibles   = habitaciones.filter(estado='DISPONIBLE').count()
    ocupadas      = habitaciones.filter(estado='OCUPADA').count()
    limpieza      = habitaciones.filter(estado='LIMPIEZA').count()
    mantenimiento = habitaciones.filter(estado='MANTENIMIENTO').count()

    estados_data   = [disponibles, ocupadas, limpieza, mantenimiento]
    estados_labels = ['Disponible', 'Ocupada', 'En Limpieza', 'Mantenimiento']
    tipo_labels, tipo_ingresos = GerenciaService.get_grafico_tipos()

    return render(request, 'gerencia/dashboard.html', {
        **kpis,
        'meses_labels':    json.dumps(meses_labels),
        'meses_ocupacion': json.dumps(meses_ocupacion),
        'estados_data':    json.dumps(estados_data),
        'estados_labels':  json.dumps(estados_labels),
        'tipo_labels':     json.dumps(tipo_labels),
        'tipo_ingresos':   json.dumps(tipo_ingresos),
    })


@login_required
@rol_requerido('admin')
def ocupacion(request):
    from apps.reservas.models import Reserva
    from django.db.models import Count, Sum
    from django.db.models.functions import TruncMonth

    datos_qs = Reserva.objects.filter(
        estado__in=['CONFIRMADA', 'CHECKIN', 'CHECKOUT']
    ).annotate(mes=TruncMonth('creado_en')).values('mes').annotate(
        reservas=Count('id'),
        ingresos=Sum('precio_total'),
    ).order_by('-mes')[:12]

    datos = []
    for d in datos_qs:
        noches = sum(
            r.num_noches for r in Reserva.objects.filter(
                estado__in=['CONFIRMADA', 'CHECKIN', 'CHECKOUT'],
                creado_en__month=d['mes'].month,
                creado_en__year=d['mes'].year,
            )
        )
        datos.append({
            'mes':      d['mes'].strftime('%B %Y'),
            'reservas': d['reservas'],
            'noches':   noches,
            'ingresos': d['ingresos'] or 0,
        })

    return render(request, 'gerencia/ocupacion.html', {'datos': datos})


@login_required
@rol_requerido('admin')
def usuarios(request):
    from apps.usuarios.models import Usuario
    usuarios = Usuario.objects.exclude(
        rol='cliente'
    ).order_by('rol', 'username')
    return render(request, 'gerencia/usuarios.html', {'usuarios': usuarios})


@login_required
@rol_requerido('admin')
def reportes(request):
    from datetime import date

    hoy          = timezone.now().date()
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin    = request.GET.get('fecha_fin')

    if fecha_inicio and fecha_fin:
        fi = date.fromisoformat(fecha_inicio)
        ff = date.fromisoformat(fecha_fin)
    else:
        fi = hoy.replace(day=1)
        ff = hoy

    resumen, por_tipo = GerenciaService.get_reporte_periodo(fi, ff)

    return render(request, 'gerencia/reportes.html', {
        'resumen':      resumen,
        'por_tipo':     por_tipo,
        'mes':          f'{fi.strftime("%d/%m/%Y")} — {ff.strftime("%d/%m/%Y")}',
        'fecha_inicio': str(fi),
        'fecha_fin':    str(ff),
    })


@login_required
@rol_requerido('admin')
def exportar_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from apps.reservas.models import Reserva
    from django.db.models import Sum
    from datetime import date

    hoy          = timezone.now().date()
    fecha_inicio = request.GET.get('fecha_inicio', str(hoy.replace(day=1)))
    fecha_fin    = request.GET.get('fecha_fin',    str(hoy))
    fi = date.fromisoformat(fecha_inicio)
    ff = date.fromisoformat(fecha_fin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte de Reservas'

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill('solid', fgColor='2D4A3E')
    center_align = Alignment(horizontal='center')

    headers = ['#', 'Huésped', 'DNI', 'Tipo Habitación', 'Entrada',
               'Salida', 'Noches', 'Total S/', 'Estado', 'Origen']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    reservas = Reserva.objects.filter(
        creado_en__date__gte=fi,
        creado_en__date__lte=ff,
    ).select_related('huesped', 'tipo_habitacion').order_by('-creado_en')

    for row, r in enumerate(reservas, 2):
        ws.cell(row=row, column=1,  value=r.pk)
        ws.cell(row=row, column=2,  value=r.huesped.nombre_completo)
        ws.cell(row=row, column=3,  value=r.huesped.num_doc)
        ws.cell(row=row, column=4,  value=r.tipo_habitacion.nombre)
        ws.cell(row=row, column=5,  value=str(r.fecha_entrada))
        ws.cell(row=row, column=6,  value=str(r.fecha_salida))
        ws.cell(row=row, column=7,  value=r.num_noches)
        ws.cell(row=row, column=8,  value=float(r.precio_total))
        ws.cell(row=row, column=9,  value=r.get_estado_display())
        ws.cell(row=row, column=10, value=r.get_origen_display())

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    total_row = reservas.count() + 2
    ws.cell(row=total_row, column=7, value='TOTAL').font = Font(bold=True)
    total = reservas.aggregate(t=Sum('precio_total'))['t'] or 0
    ws.cell(row=total_row, column=8, value=float(total)).font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_{fecha_inicio}_{fecha_fin}.xlsx"'
    wb.save(response)
    return response


@login_required
@rol_requerido('admin')
def exportar_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    from django.http import HttpResponse
    from apps.reservas.models import Reserva
    from django.db.models import Sum
    from datetime import date
    import io

    hoy          = timezone.now().date()
    fecha_inicio = request.GET.get('fecha_inicio', str(hoy.replace(day=1)))
    fecha_fin    = request.GET.get('fecha_fin',    str(hoy))
    fi = date.fromisoformat(fecha_inicio)
    ff = date.fromisoformat(fecha_fin)

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
    styles   = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(
        f'Reporte de Reservas — Hotel Tumán ({fi.strftime("%d/%m/%Y")} → {ff.strftime("%d/%m/%Y")})',
        styles['Title']
    ))
    elements.append(Spacer(1, 0.5*cm))

    headers = ['#', 'Huésped', 'DNI', 'Tipo', 'Entrada', 'Salida',
               'Noches', 'Total S/', 'Estado']
    data    = [headers]

    reservas = Reserva.objects.filter(
        creado_en__date__gte=fi,
        creado_en__date__lte=ff,
    ).select_related('huesped', 'tipo_habitacion').order_by('-creado_en')

    for r in reservas:
        data.append([
            str(r.pk),
            r.huesped.nombre_completo,
            r.huesped.num_doc,
            r.tipo_habitacion.nombre,
            str(r.fecha_entrada),
            str(r.fecha_salida),
            str(r.num_noches),
            f'S/ {r.precio_total}',
            r.get_estado_display(),
        ])

    total = reservas.aggregate(t=Sum('precio_total'))['t'] or 0
    data.append(['', '', '', '', '', 'TOTAL',
                 str(reservas.count()), f'S/ {total}', ''])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0),  (-1, 0),  colors.HexColor('#2D4A3E')),
        ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0),  (-1, 0),  9),
        ('ALIGN',         (0, 0),  (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, colors.HexColor('#F5F0E8')]),
        ('FONTSIZE',      (0, 1),  (-1, -1), 8),
        ('GRID',          (0, 0),  (-1, -1), 0.5, colors.HexColor('#E5DDD0')),
        ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND',    (0, -1), (-1, -1), colors.HexColor('#C4A882')),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{fecha_inicio}_{fecha_fin}.pdf"'
    return response